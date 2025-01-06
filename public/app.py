import sys
import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from PyQt6.QtWidgets import (
    QApplication,
    QMainWindow,
    QListWidgetItem,
    QLabel,
)
from PyQt6.QtCore import QThread, pyqtSignal, QTimer, Qt
from PyQt6.QtGui import QPixmap, QIcon
from design import Ui_MainWindow
import os


class YouTubeSongManager(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        # Connect buttons and enter key to search
        self.ui.search_button.clicked.connect(self.start_search)
        self.ui.add_to_excel_button.clicked.connect(self.add_to_excel)

        # Connect enter key to search method
        self.ui.search_input.returnPressed.connect(self.start_search)

        # List to store search results
        self.search_results = []

        # Initialize the custom spinner (QPainter) for the search button
        self.spinner_angle = 0  # Starting angle for the spinner
        self.spinner_timer = QTimer(self)
        self.spinner_timer.timeout.connect(self.update_spinner)

        # Set focus on search input field once the window opens
        self.ui.search_input.setFocus()

        # Initialize the thread as None (to keep track of it)
        self.search_thread = None

        # Label for displaying messages
        self.message_label = QLabel("")
        self.message_label.setStyleSheet("font-size: 16px; color: #FFEB3B;")
        self.ui.layout.insertWidget(0, self.message_label)

    def start_search(self):
        # Disable the button during search
        self.ui.search_button.setEnabled(False)

        # Start the spinner animation
        self.spinner_timer.start(50)  # Update spinner every 50ms

        # Run the search process in a separate thread to avoid freezing the UI
        if self.search_thread is not None and self.search_thread.isRunning():
            self.search_thread.terminate()  # Terminate any running thread if it's still active

        self.search_thread = SearchThread(self.ui.search_input.text())
        self.search_thread.results_signal.connect(self.on_search_results)
        self.search_thread.finished.connect(
            self.on_search_finished
        )  # Connect finished signals
        self.search_thread.start()

    def update_spinner(self):
        # Update the angle of the spinner
        self.spinner_angle += 10
        if self.spinner_angle >= 360:
            self.spinner_angle = 0

        # Force a repaint of the search button to show the spinner
        self.ui.search_button.update()

    def on_search_results(self, results):
        # Enable the search button back
        self.ui.search_button.setEnabled(True)

        # Stop the spinner animation
        self.spinner_timer.stop()

        # Clear previous results
        self.ui.results_list.clear()
        self.search_results = results

        # Display the search results with thumbnails
        for result in results:
            list_item = QListWidgetItem(result["title"])
            thumbnail = QPixmap()
            thumbnail.loadFromData(requests.get(result["thumbnail"]).content)
            list_item.setIcon(
                QIcon(thumbnail.scaled(150, 150, Qt.AspectRatioMode.KeepAspectRatio))
            )
            self.ui.results_list.addItem(list_item)

    def on_search_finished(self):
        # This function is called when the search thread finishes
        print("Search completed")

    def add_to_excel(self):
        selected_items = self.ui.results_list.selectedItems()
        if not selected_items:
            self.message_label.setText("Please select at least one song!")
            return

        # Use the existing Excel file
        excel_file = os.path.join(
            os.path.dirname(__file__), "songs.xlsx"
        )  # Use the full path to the file
        try:
            workbook = load_workbook(excel_file)
            sheet = workbook["Active"]
        except FileNotFoundError:
            self.message_label.setText("Excel file not found! Please ensure it exists.")
            return

        # Find the columns for "Title" and "YouTube Link"
        title_col = None
        link_col = None
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            if cell_value == "Title":
                title_col = col
            elif cell_value == "YouTube Link":
                link_col = col

        if title_col is None or link_col is None:
            self.message_label.setText(
                "Could not find 'Title' or 'YouTube Link' columns in the Excel sheet."
            )
            return

        # Add selected items to the DataFrame
        new_rows = []
        for item in selected_items:
            for result in self.search_results:
                if result["title"] == item.text():
                    new_rows.append([result["title"], result["url"]])
                    break

        # Insert new rows at the top using pandas and openpyxl
        try:
            df = pd.read_excel(excel_file, sheet_name="Active")
            new_df = pd.DataFrame(new_rows, columns=["Title", "YouTube Link"])
            df = pd.concat([new_df, df], ignore_index=True)

            # Clear the existing sheet
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                for cell in row:
                    cell.value = None

            # Write the updated DataFrame to the sheet
            for r_idx, row in enumerate(df.itertuples(index=False, name=None), 2):
                sheet.cell(row=r_idx, column=title_col, value=row[0])
                sheet.cell(row=r_idx, column=link_col, value=row[1])

            workbook.save(excel_file)
            self.ui.add_to_excel_button.setText("Songs added to Excel successfully!")
            QTimer.singleShot(
                2000,
                lambda: self.ui.add_to_excel_button.setText("Add Selected to Excel"),
            )
        except Exception as e:
            self.message_label.setText(f"Failed to save to Excel: {e}")


class SearchThread(QThread):
    results_signal = pyqtSignal(list)

    def __init__(self, search_query):
        super().__init__()
        self.search_query = search_query

    def run(self):
        API_KEY = "AIzaSyDmXg_MlBEvUb3oAtMpj-fi4Fet80b21fM"
        url = f"https://www.googleapis.com/youtube/v3/search?part=snippet&q={self.search_query}&type=video&maxResults=25&key={API_KEY}"

        try:
            response = requests.get(url)
            response.raise_for_status()
            data = response.json()

            results = []
            for item in data.get("items", []):
                video_title = item["snippet"]["title"]
                video_id = item["id"]["videoId"]
                video_url = f"https://www.youtube.com/watch?v={video_id}"
                thumbnail_url = item["snippet"]["thumbnails"]["high"]["url"]

                results.append(
                    {"title": video_title, "url": video_url, "thumbnail": thumbnail_url}
                )

            # Emit results signal to update the UI
            self.results_signal.emit(results)

        except requests.exceptions.RequestException as e:
            self.results_signal.emit([])  # If error, return empty list


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = YouTubeSongManager()
    window.show()
    sys.exit(app.exec())
