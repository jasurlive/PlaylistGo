import sys
import os
import requests
import pandas as pd
from openpyxl import load_workbook
from PyQt6.QtWidgets import QApplication, QMainWindow, QListWidgetItem, QLabel
from PyQt6.QtCore import QThread, pyqtSignal, QTimer, Qt
from PyQt6.QtGui import QPixmap, QIcon
from design import Ui_MainWindow
from dotenv import load_dotenv
import qtawesome as qta
import html  # <--- NEW: for HTML entity decoding

load_dotenv()


def normalize_text(value):
    if value is None:
        return ""
    return html.unescape(str(value))  # decode HTML entities permanently


class YouTubeSongManager(QMainWindow):
    def __init__(self):
        super().__init__()

        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        self.ui.search_button.clicked.connect(self.start_search)
        self.ui.add_to_excel_button.clicked.connect(self.add_to_excel)
        self.ui.search_input.returnPressed.connect(self.start_search)

        self.search_results = []

        self.search_thread = None

        self.message_label = QLabel("")
        self.message_label.setTextFormat(Qt.TextFormat.PlainText)
        self.message_label.setStyleSheet("font-size: 16px; color: #FFEB3B;")
        self.ui.layout.insertWidget(0, self.message_label)

        self.default_search_icon = self.ui.search_button.icon()
        self.default_search_text = self.ui.search_button.text()

        self.spinner_icon = qta.icon(
            "fa5s.spinner",
            animation=qta.Spin(self.ui.search_button),
        )

        self.ui.search_input.setFocus()

    def start_search(self):
        self.ui.search_button.setEnabled(False)
        self.ui.search_button.setText("")
        self.ui.search_button.setIcon(self.spinner_icon)

        if self.search_thread and self.search_thread.isRunning():
            self.search_thread.quit()
            self.search_thread.wait()

        query = normalize_text(self.ui.search_input.text())
        self.search_thread = SearchThread(query)
        self.search_thread.results_signal.connect(self.on_search_results)
        self.search_thread.finished.connect(self.on_search_finished)
        self.search_thread.start()

    def on_search_results(self, results):
        self.ui.search_button.setEnabled(True)
        self.ui.search_button.setIcon(self.default_search_icon)
        self.ui.search_button.setText(self.default_search_text)

        self.ui.results_list.clear()
        self.search_results = results

        for result in results:
            title = normalize_text(result["title"])
            list_item = QListWidgetItem(title)

            thumbnail = QPixmap()
            thumbnail.loadFromData(requests.get(result["thumbnail"]).content)
            list_item.setIcon(
                QIcon(
                    thumbnail.scaled(
                        150,
                        150,
                        Qt.AspectRatioMode.KeepAspectRatio,
                        Qt.TransformationMode.SmoothTransformation,
                    )
                )
            )

            self.ui.results_list.addItem(list_item)

    def on_search_finished(self):
        pass

    def add_to_excel(self):
        selected_items = self.ui.results_list.selectedItems()
        if not selected_items:
            self.message_label.setText("Please select at least one song!")
            return

        excel_file = os.path.join(os.path.dirname(__file__), "songs.xlsx")

        try:
            workbook = load_workbook(excel_file)
            sheet = workbook["Active"]
        except FileNotFoundError:
            self.message_label.setText("Excel file not found! Please ensure it exists.")
            return

        title_col = None
        link_col = None

        for col in range(1, sheet.max_column + 1):
            header = sheet.cell(row=1, column=col).value
            if header == "Title":
                title_col = col
            elif header == "YouTube Link":
                link_col = col

        if title_col is None or link_col is None:
            self.message_label.setText(
                "Could not find 'Title' or 'YouTube Link' columns in the Excel sheet."
            )
            return

        new_rows = []
        for item in selected_items:
            item_title = normalize_text(item.text())
            for result in self.search_results:
                if normalize_text(result["title"]) == item_title:
                    new_rows.append(
                        [
                            normalize_text(result["title"]),
                            normalize_text(result["url"]),
                        ]
                    )
                    break

        try:
            existing_df = pd.read_excel(excel_file, sheet_name="Active", dtype=str)

            new_df = pd.DataFrame(new_rows, columns=["Title", "YouTube Link"])
            final_df = pd.concat([new_df, existing_df], ignore_index=True)

            if sheet.max_row > 1:
                sheet.delete_rows(2, sheet.max_row - 1)

            for r_idx, row in enumerate(
                final_df.itertuples(index=False, name=None), start=2
            ):
                sheet.cell(row=r_idx, column=title_col, value=row[0])
                sheet.cell(row=r_idx, column=link_col, value=row[1])

            workbook.save(excel_file)

            self.ui.add_to_excel_button.setText("Songs added to Excel successfully!")
            QTimer.singleShot(
                2000,
                lambda: self.ui.add_to_excel_button.setText("Add Selected to Excel"),
            )

        except Exception as e:
            self.message_label.setText(f"Failed to save to Excel: {e}")


class SearchThread(QThread):
    results_signal = pyqtSignal(list)

    def __init__(self, search_query):
        super().__init__()
        self.search_query = normalize_text(search_query)

    def run(self):
        API_KEY = os.getenv("APP_YT_KEY")
        url = (
            "https://www.googleapis.com/youtube/v3/search"
            f"?part=snippet&q={self.search_query}"
            "&type=video&maxResults=20"
            f"&key={API_KEY}"
        )

        try:
            response = requests.get(url, timeout=10)
            response.raise_for_status()
            data = response.json()

            results = []
            for item in data.get("items", []):
                if "videoId" not in item.get("id", {}):
                    continue

                title = normalize_text(item["snippet"]["title"])
                video_id = item["id"]["videoId"]
                video_url = f"https://www.youtube.com/watch?v={video_id}"
                thumbnail_url = item["snippet"]["thumbnails"]["high"]["url"]

                results.append(
                    {
                        "title": title,
                        "url": normalize_text(video_url),
                        "thumbnail": thumbnail_url,
                    }
                )

            self.results_signal.emit(results)

        except requests.exceptions.RequestException:
            self.results_signal.emit([])


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = YouTubeSongManager()
    window.show()
    sys.exit(app.exec())
