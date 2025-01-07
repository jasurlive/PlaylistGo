import sys
import os
import pandas as pd
from openpyxl import load_workbook
from PyQt6.QtWidgets import (
    QApplication,
    QMainWindow,
    QListWidget,
    QListWidgetItem,
    QVBoxLayout,
    QPushButton,
    QWidget,
    QLabel,
    QAbstractItemView,
    QLineEdit,
    QDialog,
    QFormLayout,
    QDialogButtonBox,
    QHBoxLayout,
    QMenu,
    QStyle,
    QStyledItemDelegate,
    QStyleOptionButton,
    QStyleOptionViewItem,
    QToolButton,
)
from PyQt6.QtCore import Qt, QTimer, QRect, QSize
from PyQt6.QtGui import QCursor, QIcon
import qtawesome as qta


class EditSongDialog(QDialog):
    def __init__(self, title, link, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Edit Song")
        self.setFixedSize(500, 300)  # Bigger window dimensions

        self.layout = QFormLayout(self)

        self.title_edit = QLineEdit(self)
        self.title_edit.setText(title)
        self.title_edit.setStyleSheet(
            """
            QLineEdit {
                padding: 10px;
                font-size: 16px;
            }
            QLineEdit:hover {
                border: 1px solid #3c40c6;
            }
        """
        )
        self.layout.addRow("Title:", self.title_edit)

        self.link_edit = QLineEdit(self)
        self.link_edit.setText(link)
        self.link_edit.setStyleSheet(
            """
            QLineEdit {
                padding: 10px;
                font-size: 16px;
            }
            QLineEdit:hover {
                border: 1px solid #3c40c6;
            }
        """
        )
        self.layout.addRow("YouTube Link:", self.link_edit)

        self.buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel,
            self,
        )
        self.buttons.accepted.connect(self.accept)
        self.buttons.rejected.connect(self.reject)
        self.layout.addWidget(self.buttons)

        self.setStyleSheet(
            """
            QDialogButtonBox {
                padding: 10px;
            }
            QDialogButtonBox:hover {
                background-color: #3c40c6;
                color: #ffffff;
            }
        """
        )

    def get_data(self):
        return self.title_edit.text(), self.link_edit.text()

    def showEvent(self, event):
        super().showEvent(event)
        screen_geometry = QApplication.primaryScreen().geometry()
        self.move(screen_geometry.center() - self.rect().center())


class EditButtonDelegate(QStyledItemDelegate):
    def paint(self, painter, option, index):
        super().paint(painter, option, index)
        if option.state & QStyle.StateFlag.State_MouseOver:
            edit_button = QStyleOptionButton()
            edit_button.rect = QRect(
                option.rect.right() - 60, option.rect.top(), 30, option.rect.height()
            )
            edit_button.icon = qta.icon("fa.edit")
            edit_button.iconSize = QSize(20, 20)
            edit_button.state = QStyle.StateFlag.State_Enabled
            QApplication.style().drawControl(
                QStyle.ControlElement.CE_PushButton, edit_button, painter
            )

            delete_button = QStyleOptionButton()
            delete_button.rect = QRect(
                option.rect.right() - 30, option.rect.top(), 30, option.rect.height()
            )
            delete_button.icon = qta.icon("fa.trash")
            delete_button.iconSize = QSize(20, 20)
            delete_button.state = QStyle.StateFlag.State_Enabled
            QApplication.style().drawControl(
                QStyle.ControlElement.CE_PushButton, delete_button, painter
            )

    def editorEvent(self, event, model, option, index):
        if event.type() == event.Type.MouseButtonPress:
            if option.rect.right() - 60 <= event.pos().x() <= option.rect.right() - 30:
                title = index.data(Qt.ItemDataRole.DisplayRole)
                link = index.data(Qt.ItemDataRole.UserRole)
                dialog = EditSongDialog(title, link)
                if dialog.exec() == QDialog.DialogCode.Accepted:
                    new_title, new_link = dialog.get_data()
                    model.setData(index, new_title, Qt.ItemDataRole.DisplayRole)
                    model.setData(index, new_link, Qt.ItemDataRole.UserRole)
                return True
            elif option.rect.right() - 30 <= event.pos().x() <= option.rect.right():
                title = index.data(Qt.ItemDataRole.DisplayRole)
                self.parent().delete_song_by_title(title)
                return True
        elif event.type() == event.Type.MouseMove:
            if option.rect.right() - 60 <= event.pos().x() <= option.rect.right():
                QApplication.setOverrideCursor(Qt.CursorShape.PointingHandCursor)
            else:
                QApplication.restoreOverrideCursor()
        return super().editorEvent(event, model, option, index)


class SongReorderWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Reorder Songs")
        self.setGeometry(100, 100, 800, 600)  # Larger window dimensions

        # Main layout
        self.main_layout = QVBoxLayout()

        # List widget for reordering songs
        self.reorder_list = QListWidget()
        self.reorder_list.setDragDropMode(QAbstractItemView.DragDropMode.InternalMove)
        self.reorder_list.setItemDelegate(EditButtonDelegate(self))
        self.reorder_list.setStyleSheet(
            """
            QListWidget {
                background-color: #1e272e;
                color: #d2dae2;
                padding: 10px;
                border-radius: 10px;
            }
            QListWidget::item {
                padding: 10px;
                margin: 5px;
                border-radius: 5px;
                background-color: #485460;
            }
            QListWidget::item:hover {
                background-color: #006435;
            }
            QListWidget::item:selected {
                background-color: #ff60c0;
                color: #ffffff;
            }
            QToolButton {
                border: none;
                background: transparent;
            }
            QToolButton:hover {
                cursor: pointer;
                transform: scale(1.2);
            }
        """
        )
        self.reorder_list.viewport().setCursor(QCursor(Qt.CursorShape.OpenHandCursor))
        self.reorder_list.viewport().installEventFilter(self)
        self.main_layout.addWidget(self.reorder_list)

        # Button to save the new order to Excel
        self.save_order_button = QPushButton("Save Order to Excel")
        self.save_order_button.setStyleSheet(
            """
            QPushButton {
                background-color: #3c40c6;
                color: #ffffff;
                padding: 10px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #575fcf;
            }
        """
        )
        self.save_order_button.clicked.connect(self.save_order_to_excel)
        self.main_layout.addWidget(self.save_order_button)

        # Label for displaying messages
        self.message_label = QLabel("")
        self.message_label.setStyleSheet(
            "font-size: 16px; color: #ffd32a; padding: 10px;"
        )
        self.main_layout.addWidget(self.message_label)

        # Set the main layout
        container = QWidget()
        container.setLayout(self.main_layout)
        self.setCentralWidget(container)

        # Load songs from Excel
        self.load_songs_from_excel()

    def load_songs_from_excel(self):
        excel_file = os.path.join(os.path.dirname(__file__), "songs.xlsx")
        try:
            df = pd.read_excel(excel_file, sheet_name="Active")
            for title, link in zip(df["Title"], df["YouTube Link"]):
                list_item = QListWidgetItem(title)
                list_item.setData(Qt.ItemDataRole.UserRole, link)
                self.reorder_list.addItem(list_item)
        except FileNotFoundError:
            self.message_label.setText("Excel file not found! Please ensure it exists.")
        except Exception as e:
            self.message_label.setText(f"Failed to load songs: {e}")

    def save_order_to_excel(self):
        # Get the new order of songs from the reorder list
        new_order = [
            (
                self.reorder_list.item(i).text(),
                self.reorder_list.item(i).data(Qt.ItemDataRole.UserRole),
            )
            for i in range(self.reorder_list.count())
        ]

        # Use the existing Excel file
        excel_file = os.path.join(os.path.dirname(__file__), "songs.xlsx")
        try:
            workbook = load_workbook(excel_file)
            sheet = workbook["Active"]
        except FileNotFoundError:
            self.message_label.setText("Excel file not found! Please ensure it exists.")
            return

        # Find the columns for "Title" and "YouTube Link"
        title_col = None
        link_col = None
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            if cell_value == "Title":
                title_col = col
            elif cell_value == "YouTube Link":
                link_col = col

        if title_col is None or link_col is None:
            self.message_label.setText(
                "Could not find 'Title' or 'YouTube Link' columns in the Excel sheet."
            )
            return

        # Update the sheet with the new order
        for row_idx, (title, link) in enumerate(new_order, start=2):
            sheet.cell(row=row_idx, column=title_col, value=title)
            sheet.cell(row=row_idx, column=link_col, value=link)

        try:
            workbook.save(excel_file)
            self.message_label.setText("Order saved to Excel successfully!")
        except Exception as e:
            self.message_label.setText(f"Failed to save to Excel: {e}")

    def delete_selected_song(self):
        # This method can be removed if not used elsewhere
        pass

    def delete_song_by_title(self, title_to_delete):
        # Use the existing Excel file
        excel_file = os.path.join(os.path.dirname(__file__), "songs.xlsx")
        sheet_name = "Active"
        title_col_name = "Title"
        link_col_name = "YouTube Link"

        try:
            workbook = load_workbook(excel_file)
            sheet = workbook[sheet_name]
        except FileNotFoundError:
            self.message_label.setText("Excel file not found! Please ensure it exists.")
            return

        # Find the columns for "Title" and "YouTube Link"
        title_col = None
        link_col = None
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            if cell_value == title_col_name:
                title_col = col
            elif cell_value == link_col_name:
                link_col = col

        if title_col is None or link_col is None:
            self.message_label.setText(
                "Could not find 'Title' or 'YouTube Link' columns in the Excel sheet."
            )
            return

        # Find the row to delete
        row_to_delete = None
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=title_col).value == title_to_delete:
                row_to_delete = row
                break

        if row_to_delete is None:
            self.message_label.setText("Could not find the song in the Excel sheet.")
            return

        # Delete the row and shift remaining rows up
        sheet.delete_rows(row_to_delete)

        # Shift remaining rows up
        for row in range(row_to_delete, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row=row, column=col).value = sheet.cell(
                    row=row + 1, column=col
                ).value

        # Clear the last row
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=sheet.max_row, column=col).value = None

        try:
            workbook.save(excel_file)
            self.message_label.setText("Song deleted successfully!")
            for i in range(self.reorder_list.count()):
                if self.reorder_list.item(i).text() == title_to_delete:
                    self.reorder_list.takeItem(i)
                    break
        except Exception as e:
            self.message_label.setText(f"Failed to delete the song: {e}")

    def eventFilter(self, source, event):
        if (
            event.type() == event.Type.MouseButtonPress
            and source is self.reorder_list.viewport()
        ):
            source.setCursor(QCursor(Qt.CursorShape.ClosedHandCursor))
        elif (
            event.type() == event.Type.MouseButtonRelease
            and source is self.reorder_list.viewport()
        ):
            source.setCursor(QCursor(Qt.CursorShape.OpenHandCursor))
        elif (
            event.type() == event.Type.MouseMove
            and source is self.reorder_list.viewport()
        ):
            if event.buttons() == Qt.MouseButton.LeftButton:
                source.setCursor(QCursor(Qt.CursorShape.ClosedHandCursor))
            else:
                source.setCursor(QCursor(Qt.CursorShape.OpenHandCursor))
        return super().eventFilter(source, event)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = SongReorderWindow()
    window.show()
    sys.exit(app.exec())
